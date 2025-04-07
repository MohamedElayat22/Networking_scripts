import pandas as pd
import requests
from datetime import datetime
import sys
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill

# Function to check IP reputation via AbuseIPDB API
def check_ip_reputation(ip, api_key):
    url = "https://api.abuseipdb.com/api/v2/check"
    headers = {
        'Accept': 'application/json',
        'Key': api_key
    }
    params = {
        'ipAddress': ip,
        'maxAgeInDays': '90'
    }
    
    response = requests.get(url, headers=headers, params=params)
    
    if response.status_code != 200:
        print(f"Error fetching data for {ip}: {response.json()}")
        return {"data": {"abuseConfidenceScore": 0, "reputation": "No data"}}

    return response.json()

# Function to apply conditional formatting to the Excel sheet
def apply_conditional_formatting(worksheet, abuse_confidence_score_col):
    for row in range(2, worksheet.max_row + 1):
        score = worksheet.cell(row=row, column=abuse_confidence_score_col).value
        if score is not None:
            if score > 80:
                worksheet.cell(row=row, column=abuse_confidence_score_col).fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
            elif score > 50:
                worksheet.cell(row=row, column=abuse_confidence_score_col).fill = PatternFill(start_color="FFAAAA", end_color="FFAAAA", fill_type="solid")
            elif score > 49:
                worksheet.cell(row=row, column=abuse_confidence_score_col).fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            else:
                worksheet.cell(row=row, column=abuse_confidence_score_col).fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")

# Main function
def main():
    if len(sys.argv) < 2:
        print("Usage: python script.py <input_excel_file>")
        sys.exit(1)

    file_path = sys.argv[1]
    api_key = 'your_real_api_key_here'  # Replace with your real API key

    # Loading the input Excel file
    df = pd.read_excel(file_path, engine='openpyxl')

    ips = df['source_ip'].dropna().tolist()
    ip_counts = pd.Series(ips).value_counts()
    action_counts = df.groupby('source_ip')['request_status'].value_counts().unstack(fill_value=0)

    sorted_ips = ip_counts.sort_values(ascending=False)

    findings = pd.DataFrame(sorted_ips).reset_index()
    findings.columns = ['source_ip', 'count']

    action_summary = action_counts.reset_index()
    findings = pd.merge(findings, action_summary, on='source_ip', how='left')

    findings['abuse_confidence_score'] = None
    findings['reputation'] = None

    for ip in sorted_ips.index:
        reputation_data = check_ip_reputation(ip, api_key)
        findings.loc[findings['source_ip'] == ip, 'abuse_confidence_score'] = reputation_data.get("data", {}).get("abuseConfidenceScore", 0)
        findings.loc[findings['source_ip'] == ip, 'reputation'] = reputation_data

    datetime_str = datetime.now().strftime("%Y-%m-%d_%H-%M-%S")
    output_filename = f"findings_{datetime_str}.xlsx"
    findings.to_excel(output_filename, index=False, engine='openpyxl')

    # Open the workbook and get the worksheet
    wb = load_workbook(output_filename)
    ws = wb.active

    abuse_confidence_score_col = findings.columns.get_loc('abuse_confidence_score') + 1
    
    apply_conditional_formatting(ws, abuse_confidence_score_col)

    wb.save(output_filename)
    print(f"\nFindings have been saved to: {output_filename}")

if __name__ == "__main__":
    main()
